import openpyxl
class Openxl:
    #  path为要加载的文件路径  sheet_name为要加载的sheet名称
    def __init__(self,path,sheet_name):
        self.path=path
        self.sheet_name=sheet_name
        wb=openpyxl.load_workbook(self.path)
        self.ws=wb[sheet_name]
    #读取选定行列范围内的数据 每一行存入一个列表
    def read_xl_list(self,min_row,min_col,max_row,max_col):
        all_list=[]
        for rows in self.ws.iter_rows(min_row=min_row,min_col=min_col,max_row=max_row,max_col=max_col):
            row_list=[row.value for row in rows]
            all_list.append(row_list)
        return all_list

    '''必须读取连续列数据  且header写在第一行  默认读取数据从第二行开始
    header参数传读取列的title  类型为列表
    col_strar表示要读取的开始列  缺省参数默认从第一列开始读取
    每行按照header传值中的key存入一个字典
    '''
    def read_xl_dict(self,header,col_start=1):
        all_list=[]
        for row in range(2,self.ws.max_row+1):
            dic={}
            col=col_start
            for key in header:
                dic[key]=self.ws.cell(row,col).value
                col+=1
            all_list.append(dic)
        return all_list

if __name__ == '__main__':
    op=Openxl("../test_data/login.xlsx","user_fail")
    # re=op.read_xl_dict(["username","password","except_find_elemt"])
    re=op.read_xl_list(2,1,5,3)
    print(re)


